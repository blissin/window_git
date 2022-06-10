import cv2
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from matplotlib import colors

class cv2xls():
    def __init__(self,resizing):
        self.resizing=int(resizing)
        self.cell_size=2

        pass

    def img_2_xls(self,file_direct):
        try:
            img = cv2.imread(file_direct, 1) #1 : cv2.IMREAD_COLOR / 0 : cv2.IMREAD_GRAYSCALE / -1 : cv2.IMREAD_UNCHANGED
            h,w,color = img.shape
            img = cv2.resize(img,(int(h*self.resizing/200), int(w*self.resizing/200)), interpolation=cv2.INTER_AREA)

            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.title = "im2xls"

            cell_size=self.cell_size
            h, w, color= img.shape
            for x in range(w):
                sheet.column_dimensions[get_column_letter(x+1)].width = 0.5*cell_size
                for y in range(h):
                    b, g, r, = img[y, x]
                    # print(b,g,r)
                    color_code = colors.to_hex([r/255, g/255, b/255])[1:] #xxxxxx
                    sheet.cell(row=y + 1, column=x + 1).fill = PatternFill(
                        start_color=color_code,
                        end_color=color_code,
                        fill_type="solid"
                    )
                    sheet.row_dimensions[y].height = 3*cell_size

            wb.save("./im2xls.xlsx")
        except Exception as e:
            print("작업 중 오류가 발생하였습니다 : ", e)
            pass

if __name__ == '__main__':
    img_xls=cv2xls()
    img_xls.img_2_xls('')
